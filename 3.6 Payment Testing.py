import pandas as pd
import os
import re
import openpyxl
import numpy as np
import tkinter as tk
from tkinter import StringVar, Label, Entry, Button, W, filedialog, messagebox, Frame
from openpyxl import Workbook
from tkinter import filedialog, messagebox
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from tkinter import messagebox
from tkinter import ttk, filedialog, messagebox
from tkinter import StringVar, Label, Entry, Button, W, Frame
from openpyxl.utils import get_column_letter
import logging


class PayrollApp:
    def __init__(self, master):
        self.master = master
        master.title("DTC Payment v3.0")
        master.geometry("900x600")
        master.configure(bg="#f0f0f0")  # 设置背景色

        # 创建主框架
        self.main_frame = Frame(self.master, bg="#f0f0f0")
        self.main_frame.pack(expand=True, fill='both', padx=20, pady=20)

        # 文件路径存储
        self.timesheet_path = StringVar()
        self.beeline_path = StringVar()
        self.ncs_data_path = StringVar()
        self.output_path = StringVar()

        # 创建界面组件
        self.create_widgets()

    def create_widgets(self):
        # 标题
        Label(self.main_frame, text="DTC Payment Processing", font=('Helvetica', 18, 'bold'), bg="#f0f0f0").pack(pady=(0, 20))

        # 文件选择区域
        file_frame = Frame(self.main_frame, bg="#f0f0f0")
        file_frame.pack(fill='x', pady=10)

        Label(file_frame, text="Files Selection", font=('Helvetica', 14), bg="#f0f0f0").pack(anchor='w')

        self.create_file_selector(file_frame, "Timesheet.xlsx", self.timesheet_path)
        self.create_file_selector(file_frame, "Beeline.xlsx", self.beeline_path)
        self.create_file_selector(file_frame, "NCS Data.xlsx", self.ncs_data_path)

        # 自定义输入区域
        custom_frame = Frame(self.main_frame, bg="#f0f0f0")
        custom_frame.pack(fill='x', pady=20)
        Label(custom_frame, text="Custom Parameters", font=('Helvetica', 14), bg="#f0f0f0").pack(anchor='w')

        param_frame = Frame(custom_frame, bg="#f0f0f0")
        param_frame.pack(fill='x', pady=10)
        # Billing Month
        Label(param_frame, text="Billing Month (YYYY-mm):", bg="#f0f0f0").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.billing_month_entry = Entry(param_frame, width=15)
        self.billing_month_entry.grid(row=0, column=1, sticky='w', padx=5, pady=5)

        # Standard Working Days
        Label(param_frame, text="Standard Working Days:", bg="#f0f0f0").grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.standard_days_entry = Entry(param_frame, width=15)
        self.standard_days_entry.grid(row=1, column=1, sticky='w', padx=5, pady=5)

        # Timesheet Date Filter
        Label(param_frame, text="Timesheet Date Filter (dd/mm/yyyy):", bg="#f0f0f0").grid(row=2, column=0, sticky='w', padx=5, pady=5)
        self.timesheet_filter_entry = Entry(param_frame, width=15)
        self.timesheet_filter_entry.grid(row=2, column=1, sticky='w', padx=5, pady=5)      

        # 输出区域
        output_frame = Frame(self.main_frame, bg="#f0f0f0")
        output_frame.pack(fill='x', pady=10)

        Label(output_frame, text="Output", font=('Helvetica', 14), bg="#f0f0f0").pack(anchor='w')

        self.create_file_selector(output_frame, "Output Path:", self.output_path, is_dir=True)

        # 处理按钮
        self.process_btn = ttk.Button(self.main_frame, text="Start Process", command=self.process_files, style='Accent.TButton')
        self.process_btn.pack(pady=20)

        # 状态显示
        self.status_label = Label(self.main_frame, text="", fg="gray", bg="#f0f0f0")
        self.status_label.pack(pady=10)

    def create_file_selector(self, parent, label_text, path_var, is_dir=False):
        frame = Frame(parent, bg="#f0f0f0")
        frame.pack(fill='x', pady=5)

        Label(frame, text=label_text, bg="#f0f0f0").pack(side='left', padx=(0, 10))
        Entry(frame, textvariable=path_var, width=50).pack(side='left', expand=True, fill='x', padx=(0, 10))
        Button(frame, text="Browse", command=lambda: self.select_file(path_var, is_dir)).pack(side='left')

    def select_file(self, path_var, is_dir):
        if is_dir:
            directory = filedialog.askdirectory()
            if directory:
                path_var.set(directory)
        else:
            filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
            if filename:
                path_var.set(filename)

    def process_files(self):
        try:
            # 创建文件字典
            files = {
                'timesheet': self.timesheet_path.get(),
                'beeline': self.beeline_path.get(),
                'vendor': self.ncs_data_path.get()
                
            }

            # 检查所有文件是否都已选择
            if not all(files.values()):
                raise ValueError("请先选择所有必需的文件")

            # 获取其他输入参数
            billing_month = self.billing_month_entry.get()
            standard_days = int(self.standard_days_entry.get())
            timesheet_filter = self.timesheet_filter_entry.get()

            # 校验输入
            if not re.match(r"^\d{4}-\d{2}$", billing_month):
                raise ValueError("Billing Month 格式不正确，请使用YYYY-mm格式")
            if standard_days <= 0:
                raise ValueError("标准工作日必须为正整数")
            if not re.match(r"^\d{2}/\d{2}/\d{4}$", timesheet_filter):
                raise ValueError("Timesheet 日期过滤格式不正确，请使用 dd/mm/yyyy 格式")

            # 处理文件
            result, wb = process_attendance_payroll(
                timesheet_path=files['timesheet'],
                beeline_path=files['beeline'],
                vendor_path=files['vendor'],
                billing_month=billing_month.replace("-", ""),
                standard_days=standard_days,
                timesheet_filter=timesheet_filter
            )

            # 保存输出文件
            output_path = self.output_path.get()
            if not output_path:
                output_path = os.path.join(os.path.expanduser("~"), "Desktop", "DTC Payment Template")
            os.makedirs(output_path, exist_ok=True)
            output_file = os.path.join(output_path, f"DTC Payment Template_{self.billing_month_entry.get()}.xlsx")
            wb.save(output_file)
            self.status_label.config(text="")
            messagebox.showinfo("Finish!", f"The file has been successfully generated\n Save Path:{output_file}")

        except Exception as e:
            self.status_label.config(text="")
            messagebox.showerror("错误", str(e))

# 新增函数：为单元格添加边框
def add_border_to_cell(cell):
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    cell.border = thin_border

# 新增函数：将DataFrame写入工作表并添加边框
def write_dataframe_to_sheet(ws, df, start_row=1, start_col=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, start_col):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            add_border_to_cell(cell)


def process_attendance_payroll(
    timesheet_path,
    beeline_path,
    vendor_path,
    billing_month,
    standard_days,
    timesheet_filter
):
    try:
        logging.info("开始处理薪资数据...")
        # 读取 Beeline 文件
        beeline_df = pd.read_excel(beeline_path, header=0)
        # 读取 Timesheet 文件
        timesheet_df = pd.read_excel(timesheet_path, header=0)
        # 读取 Vendor 文件
        vendor_df = pd.read_excel(vendor_path, sheet_name='SG', header=2)
        beeline_selected = beeline_df[['Contractor 1BankID_User Name', 'Contractor Name', 'Billing PC Code', 'DBS Manager Name', 'Original Start Date', 'Contract End Date', 'Assignment Bill Rate']]

        # 处理Timesheet数据  
        timesheet_df['Timesheet Period (Start Date)'] = pd.to_datetime(timesheet_df['Timesheet Period (Start Date)'], format='%d/%m/%Y')
        #设置当前月份 
        current_month = pd.to_datetime(timesheet_filter, format='%d/%m/%Y')
        timesheet_filtered = timesheet_df[
           (timesheet_df['Timesheet Period (Start Date)'].dt.year == current_month.year) &
            (timesheet_df['Timesheet Period (Start Date)'].dt.month == current_month.month)
        ]


        # 在筛选后的数据上进行后续操作
        timesheet_filtered['1BankID'] = timesheet_filtered['Contractor 1BankID_User Name'].str.split('_').str[0].str.strip()
        timesheet_grouped = timesheet_filtered.groupby('1BankID')['Operational Timesheet Units'].sum().reset_index()

        # 定义函数来分离 Assignment Bill Rate
        def split_assignment_bill_rate(rate):
            if pd.isna(rate):
                return pd.NA, pd.NA
            rate_type = ''.join(re.findall(r'[a-zA-Z]+', str(rate)))
            bill_rate = ''.join(re.findall(r'\d+\.?\d*', str(rate)))
            return rate_type, bill_rate
        #分离 Assignment Bill Rate
        beeline_selected[['Rate Type', 'Bill Rate']] = beeline_selected['Assignment Bill Rate'].apply(split_assignment_bill_rate).apply(pd.Series)
        beeline_selected = beeline_selected.rename(columns={'Contractor 1BankID_User Name': '1BankID'})

        # 删除原始的 Assignment Bill Rate 列（如果不再需要）
        beeline_selected = beeline_selected.drop('Assignment Bill Rate', axis=1)
        # vendor_df = pd.read_excel(vendor_path, sheet_name='SG', header=2) # 删除这行

        # 生成母表
        mother_table = pd.merge(vendor_df[['Employee ID', '1BankID', 'BU', 'Assignment ID','Total amt(incl. VAT)']], 
                            beeline_selected[['1BankID', 'Contractor Name', 'Billing PC Code', 'DBS Manager Name', 'Original Start Date', 'Contract End Date', 'Rate Type', 'Bill Rate']], 
                            on='1BankID', 
                            how='left',
                            suffixes=('', '_beeline'))
        mother_table = pd.merge(mother_table, timesheet_grouped, on='1BankID', how='left')

        mother_table.rename(columns={'Operational Timesheet Units': 'Actual working days'}, inplace=True)
        mother_table['Month of billing'] = billing_month
        mother_table = mother_table.rename(columns={"Rate Type_y": "Rate Type"})


          # 填充空值并计算
        mother_table['Actual working days'] = pd.to_numeric(mother_table['Actual working days'], errors='coerce')
        mother_table['Bill Rate'] = pd.to_numeric(mother_table['Bill Rate'], errors='coerce')

        # 计算 'Cost(RMB)'
        mother_table['Cost(RMB)'] = np.where(
            mother_table['Rate Type'] == 'Monthly',
            (mother_table['Actual working days'] / standard_days) * mother_table['Bill Rate'].fillna(0),
            np.where(
                mother_table['Rate Type'] == 'Daily',
            mother_table['Actual working days'] * mother_table['Bill Rate'].fillna(0),
            0  # 如果不是 Monthly，可以设置为 0 或其他默认值
            )
        )

        # 计算 'Tax (RMB)'
        mother_table['Tax (RMB)'] = np.where(
            mother_table['Rate Type'] =='Monthly',
            mother_table['Cost(RMB)'] * 0.06,
            0
        )
        mother_table['Actual working days'] = mother_table['Actual working days'].fillna(0)  #填补值为0
        mother_table['Total (RMB)'] = (mother_table['Cost(RMB)'] + mother_table['Tax (RMB)']).round(2)


        # 重新排列列顺序
        mother_table = mother_table[[
           'Employee ID', '1BankID', 'Contractor Name', 'Billing PC Code', 'BU','DBS Manager Name', 'Original Start Date', 'Contract End Date', 'Rate Type','Bill Rate','Actual working days','Cost(RMB)','Tax (RMB)','Total (RMB)','Assignment ID','Total amt(incl. VAT)',
        ]]
        mother_table.insert(0,'Month of billing',billing_month)

        # 更改列名
        mother_table = mother_table.rename(columns={"Original Start Date": "Join Date"})
        mother_table = mother_table.rename(columns={"Contract End Date": "Exit Date"})
        mother_table = mother_table.rename(columns={"Assignment Bill Rate": "Rate per month(RMB)"})
        mother_table = mother_table.rename(columns={"Actual working days": "No. of days"})
        mother_table = mother_table.rename(columns={"Total amt(incl. VAT)": "NCS Value"})

        # 首先确保 'NCS Value' 和 'Total (RMB)' 列都是数值类型
        mother_table['NCS Value'] = pd.to_numeric(mother_table['NCS Value'], errors='coerce')
        mother_table['Total (RMB)'] = pd.to_numeric(mother_table['Total (RMB)'], errors='coerce')

        # 创建 'Validation' 列f
        mother_table['Validation'] = mother_table['NCS Value'] == mother_table['Total (RMB)']
        mother_table['Validation'] = mother_table['Validation'].map({True: 'True', False: 'False'})

        
        # 将 'Contract Start Date' 列转换为 datetime 对象
        mother_table['Join Date'] = pd.to_datetime(mother_table['Join Date'], format='%d/%m/%Y')
        mother_table['Join Date'] = mother_table['Join Date'].dt.strftime('%Y/%m/%d')

        # 首先确保 'Contract End Date' 列是 datetime 类型
        mother_table['Exit Date'] = pd.to_datetime(mother_table['Exit Date'], errors='coerce')
        mother_table['Exit Date'] = mother_table['Exit Date'].dt.date

        #Contract End Date 日期筛选
        mother_table['Exit Date'] = pd.to_datetime(mother_table['Exit Date'], dayfirst=True)
        if 'Exit Date' in mother_table.columns:
        # 提取目标年月
            target_year = int(billing_month[:4])
            target_month = int(billing_month[4:6])
        # 转换为datetime类型并筛选
        mother_table['Exit Date'] = pd.to_datetime(
            mother_table['Exit Date'],
            errors='coerce'  # 无效日期转为NaT
        )
        # 创建年月匹配条件
        mask = (
            (mother_table['Exit Date'].dt.year == target_year) & 
            (mother_table['Exit Date'].dt.month == target_month)
        )
        # 清除不符合条件的日期
        mother_table.loc[~mask, 'Exit Date'] = pd.NaT
        mother_table['Exit Date']= mother_table['Exit Date'].dt.strftime('%Y/%m/%d')

    except Exception as e:
        logging.error(f"处理过程中出现错误: {e}")
        raise

    try:
        # 创建一个新的Excel工作簿
        wb = Workbook()
        # SG PC
        ws_sg = wb.active
        ws_sg.title = "SG PC"
        write_dataframe_to_sheet(ws_sg, mother_table, start_row=4, start_col=1)

        # 插入内容到K3/L3单元格
        ws_sg['K3'] = 'Acutal Working Days'
        ws_sg['L3'] = standard_days
        add_border_to_cell(ws_sg['K3'])
        add_border_to_cell(ws_sg['L3'])

        # 为新添加的单元格添加边框
        add_border_to_cell(ws_sg['M3'])
        add_border_to_cell(ws_sg ['N3'])
        add_border_to_cell(ws_sg['O3'])
        add_border_to_cell(ws_sg['Q3'])
       
        # 使用write_dataframe_to_sheet函数写入mother_table数据
        write_dataframe_to_sheet(ws_sg, mother_table, start_row=4, start_col=1)
        # 在这里插入新的格式化代码
        # 设置A4-R4的样式
        for col in range(1, 19):  # A到R列
            cell = ws_sg.cell(row=4, column=col)
        
            # 设置字体为粗体和白色
            cell.font = Font(bold=True, color="FFFFFF")
        
            # 设置底色
            if col <= 16:  # A到P列
                cell.fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark Blue
            else:  # Q和R列
                cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange


        # 将DataFrame写入Excel，从第4行A列开始
        for r_idx, row in enumerate(dataframe_to_rows(mother_table, index=False, header=True), 4):
            for c_idx, value in enumerate(row, 1):
                ws_sg.cell(row=r_idx, column=c_idx, value=value)

        # 获取 'Validation' 列的索引 和  # 获取 '1BankID' 列的索引
        validation_col = mother_table.columns.get_loc('Validation') + 1
        bank_id_col = mother_table.columns.get_loc('1BankID') + 1
        # 创建黄色填充样式和红色字体样式
        yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
        red_font = Font(color='FF0000')
        # 用于跟踪重复的 1BankID
        bank_id_count = {}

        # 遍历所有行（跳过标题行）
        for row in range(5, ws_sg.max_row + 1):  # 从第5行开始，因为数据从第4行开始
            # 检查 'Validation' 列
            validation_cell = ws_sg.cell(row=row, column=validation_col)
            if validation_cell.value == 'False':
                validation_cell.font = red_font
            # 检查 '1BankID' 列
            bank_id_cell = ws_sg.cell(row=row, column=bank_id_col)
            bank_id = bank_id_cell.value
            if bank_id in bank_id_count:
                bank_id_cell.fill = yellow_fill
                ws_sg.cell(row=bank_id_count[bank_id], column=bank_id_col).fill = yellow_fill
            else:
                bank_id_count[bank_id] = row
         
    except Exception as e:
        logging.error(f"处理过程中出现错误: {e}")
        raise

    try:
        cn_table = pd.DataFrame()  # 初始化一个空的 DataFrame    
        # CN PC
        ws_cn = wb.create_sheet("CN PC")
        write_dataframe_to_sheet(ws_cn, cn_table, start_row=4, start_col=1)

        # 读取数据（合并vendor读取） 
        vendor_df_cn = pd.read_excel(vendor_path, sheet_name='CN', header=2)
        beeline_df = pd.read_excel(beeline_path, header=0)
        timesheet_df = pd.read_excel(timesheet_path, header=0)

         # 生成母表-CN
        vendor_selected_cn = vendor_df_cn[['Employee ID','1BankID','BU', 'Assignment ID','Total amt(incl. VAT)']]
        beeline_selected_cn = beeline_df[['Contractor 1BankID_User Name','Contractor Name', 'Billing PC Code', 'DBS Manager Name', 'Original Start Date', 'Contract End Date', 'Assignment Bill Rate']]
        
        #分离 Assignment Bill Rate
        beeline_selected_cn[['Rate Type', 'Bill Rate']] = beeline_selected_cn['Assignment Bill Rate'].apply(split_assignment_bill_rate).apply(pd.Series)
        beeline_selected_cn = beeline_selected_cn.rename(columns={'Contractor 1BankID_User Name': '1BankID'})

        # 删除原始的 Assignment Bill Rate 列（如果不再需要）
        beeline_selected_cn = beeline_selected_cn.drop('Assignment Bill Rate', axis=1)
        cn_table = pd.merge(vendor_selected_cn, beeline_selected_cn, on='1BankID', how='left')
        cn_table = pd.merge(cn_table, timesheet_grouped, on='1BankID', how='left')
        cn_table.rename(columns={'Operational Timesheet Units': 'Actual working days'}, inplace=True)
        cn_table['Month of billing'] = billing_month

        # 填充空值并计算
        cn_table['Actual working days'] = pd.to_numeric(cn_table['Actual working days'], errors='coerce')
        cn_table['Bill Rate'] = pd.to_numeric(cn_table['Bill Rate'], errors='coerce')

        # 计算 'Cost(RMB)'
        cn_table['Cost(RMB)'] = np.where(
            cn_table['Rate Type'].isin(['Monthly', 'monthly']),
            (cn_table['Actual working days'] / standard_days) * cn_table['Bill Rate'].fillna(0),
            np.where(
                cn_table['Rate Type'].isin(['Daily', 'daily']),
                cn_table['Actual working days'] * cn_table['Bill Rate'].fillna(0),
                0
            )
        )
        

        # 计算 'Tax (RMB)'
        cn_table['Tax (RMB)'] = np.where(
            cn_table['Rate Type'] =='Monthly',
            cn_table['Cost(RMB)'] * 0.06,
            0
        )
        cn_table['Actual working days'] = cn_table['Actual working days'].fillna(0)  #填补值为0
        cn_table['Total (RMB)'] = (cn_table['Cost(RMB)'] + cn_table['Tax (RMB)']).round(2)


        # 重新排列列顺序
        cn_table = cn_table[[
           'Employee ID', '1BankID', 'Contractor Name', 'Billing PC Code', 'BU','DBS Manager Name', 'Original Start Date', 'Contract End Date', 'Rate Type','Bill Rate','Actual working days','Cost(RMB)','Tax (RMB)','Total (RMB)','Assignment ID','Total amt(incl. VAT)',
        ]]
        cn_table.insert(0,'Month of billing',billing_month)

        # 更改列名
        cn_table = cn_table.rename(columns={"Original Start Date": "Join Date"})
        cn_table = cn_table.rename(columns={"Contract End Date": "Exit Date"})
        cn_table = cn_table.rename(columns={"Assignment Bill Rate": "Rate per month(RMB)"})
        cn_table = cn_table.rename(columns={"Actual working days": "No. of days"})
        cn_table = cn_table.rename(columns={"Total amt(incl. VAT)": "NCS Value"})


        # 首先确保 'NCS Value' 和 'Total (RMB)' 列都是数值类型
        cn_table['NCS Value'] = pd.to_numeric(cn_table['NCS Value'], errors='coerce')
        cn_table['Total (RMB)'] = pd.to_numeric(cn_table['Total (RMB)'], errors='coerce')

        # 创建 'Validation' 列
        cn_table['Validation'] = cn_table['NCS Value'] == cn_table['Total (RMB)']
        cn_table['Validation'] = cn_table['Validation'].map({True: 'True', False: 'False'})

        
        # 将 'Contract Start Date' 列转换为 datetime 对象
        cn_table['Join Date'] = pd.to_datetime(cn_table['Join Date'], format='%d/%m/%Y')
        cn_table['Join Date'] = cn_table['Join Date'].dt.strftime('%Y/%m/%d')

        # 首先确保 'Contract End Date' 列是 datetime 类型
        cn_table['Exit Date'] = pd.to_datetime(cn_table['Exit Date'], errors='coerce')
        cn_table['Exit Date'] = cn_table['Exit Date'].dt.date

        #Contract End Date 日期筛选
        cn_table['Exit Date'] = pd.to_datetime(cn_table['Exit Date'], dayfirst=True)
        if 'Exit Date' in cn_table.columns:
        # 提取目标年月
            target_year = int(billing_month[:4])
            target_month = int(billing_month[4:6])
        # 转换为datetime类型并筛选
        cn_table['Exit Date'] = pd.to_datetime(
            cn_table['Exit Date'],
            errors='coerce'  # 无效日期转为NaT
        )
        # 创建年月匹配条件
        mask = (
            (cn_table['Exit Date'].dt.year == target_year) & 
            (cn_table['Exit Date'].dt.month == target_month)
        )
        # 清除不符合条件的日期
        cn_table.loc[~mask, 'Exit Date'] = pd.NaT
        cn_table['Exit Date']= cn_table['Exit Date'].dt.strftime('%Y/%m/%d')

    except Exception as e:
        logging.error(f"处理过程中出现错误: {e}")
        raise

    try:
        hk_table = pd.DataFrame()  
        # HK PC
        ws_hk = wb.create_sheet("HK PC&MS")
        write_dataframe_to_sheet(ws_hk, hk_table, start_row=4, start_col=1)

        # 读取数据（合并vendor读取） 
        vendor_df_hk = pd.read_excel(vendor_path, sheet_name='HK', header=2)
        beeline_df = pd.read_excel(beeline_path, header=0)
        timesheet_df = pd.read_excel(timesheet_path, header=0)

        # 创建一个映射表
        bill_rate_mapping = {
            ('Generic', 'School Fresh', 'Monthly'): 23516,
            ('Generic', 'Software Engineer', 'Monthly'): 26818,
            ('Generic', 'Assoc Consultant', 'Monthly'): 31535,
            ('Generic', 'Associate Consultant', 'Monthly'): 31535,
            ('Generic', 'Consultant', 'Monthly'): 37047,
            ('Generic', 'Senior Consultant', 'Monthly'): 43884,
            ('Generic', 'Lead Consultant L1', 'Monthly'): 62376,
            ('Generic', 'Lead Consultant L2', 'Monthly'): 79600,
            ('Generic', 'Lead Consultant L3', 'Monthly'): 90100,
            ('Niche', 'Software Engineer', 'Monthly'): 36835,
            ('Niche', 'Associate Consultant', 'Monthly'): 41616,
            ('Niche', 'Assoc Consultant', 'Monthly'): 41616,
            ('Niche', 'Consultant', 'Monthly'): 49078,
            ('Niche', 'Senior Consultant', 'Monthly'): 57876,
            ('Niche', 'Lead Consultant L1', 'Monthly'): 87806,
            ('Niche', 'Lead Consultant L2', 'Monthly'): 96460,
            ('Niche', 'Lead Consultant L3', 'Monthly'): 106000
        }

        # 将映射表转换为DataFrame
        bill_rate_df = pd.DataFrame([(spec, role, rate_type, rate) for (spec, role, rate_type), rate in bill_rate_mapping.items()],
                                    columns=['Specialization', 'Roles', 'Rate Type', 'Bill Rate'])

        def process_job_title(job_title):
            # 使用 '-' 分割 Job Title
            parts = job_title.split('-', 1)  # 只分割一次，以防 Roles 中包含 '-'
            # 提取 Specialization（左边部分）
            specialization = parts[0].strip() if len(parts) > 0 else ''
            # 提取 Roles（右边部分，如果存在）
            roles = parts[1].strip() if len(parts) > 1 else ''
            return specialization, roles
        
        # 生成母表-HK
        vendor_selected_hk = vendor_df_hk[['Employee ID','1BankID','BU', 'Assignment ID','Total amt(incl. VAT)','Employer Name','Contract No.','Type','location','Perm/Vendor','Rate Type']]
        beeline_selected_hk = beeline_df[['Contractor 1BankID_User Name','Contractor Name', 'Billing PC Code', 'DBS Manager Name', 'Original Start Date', 'Contract End Date','Job Title']]
     
        # 处理 Job Title
        beeline_selected_hk[['Specialization', 'Roles']] = beeline_selected_hk['Job Title'].apply(process_job_title).apply(pd.Series)

        # 合并两张表，vendor+beeline， 进行Bill Rate的匹配
        hk_table = pd.merge(vendor_selected_hk, beeline_selected_hk, left_on='1BankID', right_on='Contractor 1BankID_User Name', how='left')
        hk_table = pd.merge(hk_table, bill_rate_df, on=['Specialization', 'Roles', 'Rate Type'], how='left')
        
        # 如果没有匹配到 Bill Rate，可以设置一个默认值或保留为 NaN
        hk_table['Bill Rate'] = hk_table['Bill Rate'].fillna(pd.NA)
        hk_table = pd.merge(hk_table, timesheet_grouped, on='1BankID', how='left')
        hk_table.rename(columns={'Operational Timesheet Units': 'Actual working days'}, inplace=True)
        hk_table['Month of billing'] = billing_month

        # 确保选择需要的列
        hk_table = hk_table[[
            'Employee ID', '1BankID', 'Contractor Name', 'Billing PC Code', 'DBS Manager Name',
            'Original Start Date', 'Contract End Date', 'Rate Type', 'Specialization', 'Roles',
            'Bill Rate', 'Actual working days', 'BU', 'Assignment ID', 'Total amt(incl. VAT)',
            'Employer Name', 'Contract No.', 'Type', 'location', 'Perm/Vendor', 'Month of billing'
        ]]

        hk_table.rename(columns={'Operational Timesheet Units': 'Actual working days'}, inplace=True)

        # 填充空值并计算
        hk_table['Actual working days'] = pd.to_numeric(hk_table['Actual working days'], errors='coerce')
        hk_table['Bill Rate'] = pd.to_numeric(hk_table['Bill Rate'], errors='coerce')
        hk_table['Actual working days'] = hk_table['Actual working days'].fillna(0)  #填补值为0

        def excel_column(n):
            return get_column_letter(n)

        # 在HK部分，动态计算列位置
        def get_column_positions(hk_table):
            """动态获取列位置"""
            columns = list(hk_table.columns)
            return {
                'actual_days': excel_column(columns.index('No. of days') + 1),
                'bill_rate': excel_column(columns.index('Rate per month(RMB)') + 1),
                'rate_type': excel_column(columns.index('Rate Type') + 1),
                'cost': excel_column(columns.index('Cost(RMB)') + 1),
                'tax': excel_column(columns.index('Tax (RMB)') + 1),
                'total': excel_column(columns.index('Total (RMB)') + 1),
                'ncs_value': excel_column(columns.index('NCS Value') + 1),
            }

        # 假设 'Actual working days' 在第 A 列，'Bill Rate' 在第 B 列，'Rate Type' 在第 C 列
        actual_days_col = excel_column(17)
        bill_rate_col = excel_column(16)
        rate_type_col = excel_column(15)
        cost_col = excel_column(18)
        tax_col = excel_column(19)
        total_col = excel_column(20)

        # 计算 'Cost(RMB)'
        hk_table['Cost(RMB)'] = hk_table.apply(lambda row: 
            f'=IF(OR({rate_type_col}{row.name+5}="Monthly",{rate_type_col}{row.name+5}="Daily"),'
            f'IF({rate_type_col}{row.name+5}="Monthly",'
            f'({actual_days_col}{row.name+5}/{standard_days})*{bill_rate_col}{row.name+5},'
            f'{actual_days_col}{row.name+5}*{bill_rate_col}{row.name+5}),0)', axis=1)

        # 计算 'Tax (RMB)'
        hk_table['Tax (RMB)'] = hk_table.apply(lambda row:
            f'=IF({rate_type_col}{row.name+5}="Monthly",{cost_col}{row.name+5}*0.06,0)', axis=1)
        # 计算 'Total (RMB)'
        hk_table['Total (RMB)'] = hk_table.apply(lambda row:
            f'=ROUND({cost_col}{row.name+5}+{tax_col}{row.name+5},2)', axis=1)

        # 添加Validation公式
        ncs_value_col = excel_column(22)
        total_rmb_col = excel_column(20)
        hk_table['Validation'] = hk_table.apply(lambda row:
            f'=IF({ncs_value_col}{row.name+5}={total_rmb_col}{row.name+5},"True","False")', axis=1)

        # 重新排列列顺序
        hk_table = hk_table[[
           'Employee ID', 'Contract No.','1BankID', 'Contractor Name', 'Specialization','Roles','Billing PC Code','location','Perm/Vendor','BU','Original Start Date', 'Contract End Date', 'Rate Type','Bill Rate','Actual working days','Cost(RMB)','Tax (RMB)','Total (RMB)','Assignment ID','Total amt(incl. VAT)', 'Validation'
        ]]
        hk_table.insert(0,'Month of billing',billing_month)

        # 更改列名
        hk_table = hk_table.rename(columns={"Original Start Date": "Join Date"})
        hk_table = hk_table.rename(columns={"Contract End Date": "Exit Date"})
        hk_table = hk_table.rename(columns={"Bill Rate": "Rate per month(RMB)"})
        hk_table = hk_table.rename(columns={"Actual working days": "No. of days"})
        hk_table = hk_table.rename(columns={"Total amt(incl. VAT)": "NCS Value"})

        # 将 'Contract Start Date' 列转换为 datetime 对象
        hk_table['Join Date'] = pd.to_datetime(hk_table['Join Date'], format='%d/%m/%Y')
        hk_table['Join Date'] = hk_table['Join Date'].dt.strftime('%Y/%m/%d')

        # 首先确保 'Contract End Date' 列是 datetime 类型
        hk_table['Exit Date'] = pd.to_datetime(hk_table['Exit Date'], errors='coerce')
        hk_table['Exit Date'] = hk_table['Exit Date'].dt.date

        #Contract End Date 日期筛选
        hk_table['Exit Date'] = pd.to_datetime(hk_table['Exit Date'], dayfirst=True)
        if 'Exit Date' in hk_table.columns:
            # 提取目标年月
            target_year = int(billing_month[:4])
            target_month = int(billing_month[4:6])
            # 转换为datetime类型并筛选
            hk_table['Exit Date'] = pd.to_datetime(
                hk_table['Exit Date'],
                errors='coerce'  # 无效日期转为NaT
            )
            # 创建年月匹配条件
            mask = (
                (hk_table['Exit Date'].dt.year == target_year) & 
                (hk_table['Exit Date'].dt.month == target_month)
            )
            # 清除不符合条件的日期
            hk_table.loc[~mask, 'Exit Date'] = pd.NaT

        hk_table['Exit Date'] = hk_table['Exit Date'].dt.strftime('%Y/%m/%d')

    except Exception as e:
        logging.error(f"处理过程中出现错误: {e}")
        raise

    try:
        # 插入内容到K3/L3单元格
        ws_hk['K3'] = 'Acutal Working Days'
        ws_hk['L3'] = standard_days
        add_border_to_cell(ws_hk['K3'])
        add_border_to_cell(ws_hk['L3'])

        # 为新添加的单元格添加边框
        add_border_to_cell(ws_hk['M3'])
        add_border_to_cell(ws_hk['N3'])
        add_border_to_cell(ws_hk['O3'])
        add_border_to_cell(ws_hk['Q3'])
       
        # 使用write_dataframe_to_sheet函数写入mother_table数据
        write_dataframe_to_sheet(ws_hk, hk_table, start_row=4, start_col=1)
        
        # 在这里插入新的格式化代码
        # 设置A4-R4的样式
        for col in range(1, 19):  # A到R列
            cell = ws_hk.cell(row=4, column=col)
        
            # 设置字体为粗体和白色
            cell.font = Font(bold=True, color="FFFFFF")
            # 设置底色
            if col <= 16:  # A到P列
                cell.fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark Blue
            else:  # Q和R列
                cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange

        # 将DataFrame写入Excel，从第4行A列开始
        for r_idx, row in enumerate(dataframe_to_rows(hk_table, index=False, header=True), 4):
            for c_idx, value in enumerate(row, 1):
                ws_hk.cell(row=r_idx, column=c_idx, value=value)

        # 获取 'Validation' 列的索引 和  # 获取 '1BankID' 列的索引
        validation_col = hk_table.columns.get_loc('Validation') + 1
        bank_id_col = hk_table.columns.get_loc('1BankID') + 1

        # 创建黄色填充样式和红色字体样式
        yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
        red_font = Font(color='FF0000')
        # 用于跟踪重复的 1BankID
        bank_id_count = {}

        # 遍历所有行（跳过标题行）
        for row in range(5, ws_hk.max_row + 1):  # 从第5行开始，因为数据从第4行开始
            df_row_idx = row - 5
            
            # 检查 Validation 列 - 基于实际数据比较
            if df_row_idx < len(hk_table):
                try:
                    # 获取原始数据进行比较
                    ncs_value = hk_table.iloc[df_row_idx]['NCS Value']
                    
                    # 手动计算Total (RMB)来比较（因为Excel中是公式）
                    actual_days = hk_table.iloc[df_row_idx]['No. of days']
                    bill_rate = hk_table.iloc[df_row_idx]['Rate per month(RMB)']
                    rate_type = hk_table.iloc[df_row_idx]['Rate Type']
                    
                    # 重新计算Cost和Total来比较
                    if rate_type == 'Monthly':
                        cost = (actual_days / standard_days) * bill_rate if not pd.isna(bill_rate) else 0
                        tax = cost * 0.06
                    elif rate_type == 'Daily':
                        cost = actual_days * bill_rate if not pd.isna(bill_rate) else 0
                        tax = 0
                    else:
                        cost = 0
                        tax = 0
                    
                    calculated_total = round(cost + tax, 2)
                    
                    # 比较NCS Value和计算的Total，如果不匹配则标红
                    if not pd.isna(ncs_value) and not pd.isna(calculated_total):
                        if abs(ncs_value - calculated_total) > 0.01:  # 容差0.01
                            validation_cell = ws_hk.cell(row=row, column=validation_col)
                            validation_cell.font = red_font
                            
                except (ValueError, TypeError, IndexError, KeyError):
                    # 如果计算出错，可以选择标红表示需要检查
                    validation_cell = ws_hk.cell(row=row, column=validation_col)
                    validation_cell.font = red_font
            
            # 检查 '1BankID' 列的重复
            bank_id_cell = ws_hk.cell(row=row, column=bank_id_col)
            bank_id = bank_id_cell.value
            if bank_id in bank_id_count:
                bank_id_cell.fill = yellow_fill
                ws_hk.cell(row=bank_id_count[bank_id], column=bank_id_col).fill = yellow_fill
            else:
                bank_id_count[bank_id] = row

    except Exception as e:
        logging.error(f"处理过程中出现错误: {e}")
        raise

    try:
        # 添加 Beeline 工作表
        ws_beeline = wb.create_sheet("Beeline")
        write_dataframe_to_sheet(ws_beeline, beeline_df)

        # 添加 NCS Data 工作表
        ws_ncs = wb.create_sheet("NCS Data")
        write_dataframe_to_sheet(ws_ncs, vendor_df)

        # 添加 Timesheet Details 工作表
        ws_details = wb.create_sheet("Timesheet Details")
        write_dataframe_to_sheet(ws_details, timesheet_filtered)
        for r_idx, row in enumerate(dataframe_to_rows(timesheet_filtered, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_details.cell(row=r_idx, column=c_idx, value=value)

        # 添加 Timesheet Summary 工作表
        ws_summary = wb.create_sheet("Timesheet Summary")

        # 合并 Timesheet_grouped 和 '1BankID'
        summary_data = pd.merge(timesheet_grouped, mother_table[['1BankID']], on='1BankID', how='left')
        for r_idx, row in enumerate(dataframe_to_rows(summary_data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_summary.cell(row=r_idx, column=c_idx, value=value)
        write_dataframe_to_sheet(ws_summary, summary_data)

    except Exception as e:
        logging.error(f"处理过程中出现错误: {e}")
        raise

    return mother_table, wb

def filter_exit_dates(table, billing_month):
    """统一的退出日期筛选函数"""
    if 'Exit Date' not in table.columns:
        return table
    
    # 提取目标年月
    target_year = int(billing_month[:4])
    target_month = int(billing_month[4:6])
    
    # 转换为datetime类型并筛选
    table['Exit Date'] = pd.to_datetime(table['Exit Date'], errors='coerce')
    
    # 创建年月匹配条件
    mask = (
        (table['Exit Date'].dt.year == target_year) & 
        (table['Exit Date'].dt.month == target_month)
    )
    
    # 清除不符合条件的日期
    table.loc[~mask, 'Exit Date'] = pd.NaT
    table['Exit Date'] = table['Exit Date'].dt.strftime('%Y/%m/%d')
    
    return table

def validate_and_format_cells(ws, table, validation_col, bank_id_col, standard_days):
    """统一的验证和格式化函数"""
    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
    red_font = Font(color='FF0000')
    bank_id_count = {}
    
    for row in range(5, ws.max_row + 1):
        df_row_idx = row - 5
        
        # Validation检查
        if df_row_idx < len(table):
            try:
                ncs_value = table.iloc[df_row_idx]['NCS Value']
                actual_days = table.iloc[df_row_idx]['No. of days']
                bill_rate = table.iloc[df_row_idx].get('Rate per month(RMB)', 0)
                rate_type = table.iloc[df_row_idx]['Rate Type']
                
                # 计算预期总额
                if rate_type == 'Monthly':
                    cost = (actual_days / standard_days) * bill_rate if not pd.isna(bill_rate) else 0
                    tax = cost * 0.06
                else:
                    cost = actual_days * bill_rate if not pd.isna(bill_rate) else 0
                    tax = 0
                
                calculated_total = round(cost + tax, 2)
                
                if not pd.isna(ncs_value) and abs(ncs_value - calculated_total) > 0.01:
                    validation_cell = ws.cell(row=row, column=validation_col)
                    validation_cell.font = red_font
                    
            except (ValueError, TypeError, IndexError, KeyError):
                validation_cell = ws.cell(row=row, column=validation_col)
                validation_cell.font = red_font
        
        # 1BankID重复检查
        bank_id_cell = ws.cell(row=row, column=bank_id_col)
        bank_id = bank_id_cell.value
        if bank_id in bank_id_count:
            bank_id_cell.fill = yellow_fill
            ws.cell(row=bank_id_count[bank_id], column=bank_id_col).fill = yellow_fill
        else:
            bank_id_count[bank_id] = row

if __name__ == "__main__":
    root = tk.Tk()
    app = PayrollApp(root)
    root.mainloop()
    
    # 移除或注释掉以下测试代码，避免在GUI模式下执行
    # result, wb = process_attendance_payroll(...)